#!/usr/bin/env python
# coding=utf-8
import os
import time

from datetime import datetime
from openpyxl import Workbook
from django.http import JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.core.servers.basehttp import FileWrapper
from django.views.decorators.http import require_http_methods

from business_manager.order.models import User, AddressBook, CallRecord


@require_http_methods(['GET'])
def get_address_book(request):
    """"""
    url = settings.DOMAIN + request.get_full_path()
    url = url.replace('get_address_book', 'down_address_book')
    return JsonResponse({
        'code': 0,
        'msg': '', 
        'data': {
            'url': url
        }
    })

@require_http_methods(['GET'])
def down_address_book(request):
    """下载联系人列表"""
    user = get_object_or_404(User, pk=request.GET.get('user_id'))

    wb = Workbook()
    ws = wb.create_sheet('addressbook', 0)

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S').replace('-', '').replace(' ', '').replace(':', '')
    xpath = 'addressbook_{}.xlsx'.format(now)

    ws.title = u'联系人信息'
    ws.append([u'电话号码', u'姓名', u'创建时间', u'联系次数', u'所属人'])

    addressbooks = AddressBook.objects.filter(owner=user)
    for book in addressbooks:
        try:
            created_at = ''
            if book.create_time > 0:
                created_at = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(book.create_time))
            ws.append([book.phone_number, book.name, created_at, book.call_times, user.name])
        except Exception, e:
            print 'In down_address_book -> ', e
            continue
    
    # 通话记录
    ws1 = wb.create_sheet('callrecord', 1)
    ws1.title = u'通话记录'
    ws1.append([u'电话号码', u'姓名', u'时长', u'通话类型', u'通话时间', u'所属人'])
    callrecords = CallRecord.objects.filter(owner=user)
    for record in callrecords:
        try:
            call_time = ''
            if record.call_time:
                call_time = record.call_time.strftime('%Y-%m-%d %H:%M:%S')
            ws1.append([record.phone_number, record.name, record.duration, record.get_call_type_display(), call_time, user.name])
        except Exception, e:
            print 'In down_call_records -> ', e
            continue

    wb.save(xpath)
    response = StreamingHttpResponse(FileWrapper(open(xpath), 8192), content_type='application/vnd.ms-excel')
    response['Content-Length'] = os.path.getsize(xpath)
    response['Content-Disposition'] = 'attachment; filename={}'.format(xpath)
    try:
        os.system('rm {}'.format(xpath))
    except:
        pass
    return response 


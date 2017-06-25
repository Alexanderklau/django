# -*- coding: utf-8 -*-
"""催收前后端分离接口"""
import os
import re
import json
import requests

from datetime import datetime
from openpyxl import Workbook
from xlrd import xldate_as_tuple
from collections import OrderedDict

from rest_framework.response import Response
from rest_framework import viewsets
from rest_framework import mixins, generics, pagination, filters
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.decorators import detail_route
from django.conf import settings
from django.db.models.query import QuerySet
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404
from django.db.models import Q, Sum, Count
from django.core.servers.basehttp import FileWrapper
from django.views.decorators.csrf import csrf_exempt
from dateutil.relativedelta import relativedelta
from django.core.paginator import Paginator

from business_manager.employee.models import check_employee, get_employee_platform
from business_manager.util.permission_decorator import page_permission
from business_manager.employee.models import Employee, EmployeeGroup
from business_manager.order.apply_models import (
    Apply, InfoField, InfoModule, RepairHistory
)
from business_manager.order.models import User, ContactInfo, Messagetemplate
from business_manager.review.models import CollectionRecord, DingdangRepaymentRecord, CollectionRecordTag
from business_manager.message_server.message_client import MessageClient
from business_manager.collection.models import (
    InstallmentDetailInfo, InstallmentRecord, QualityControlRecord,
    InspectionDetails
)
from business_manager.collection.report import report_collection
from business_manager.custom.general_views import get_review_staff_name
from business_manager.python_common.log_client import CommonLog as Log
from business_manager.util.tkdate import *
from business_manager.import_data.services import get_collection_type as gct
from business_manager.employee.views import check_wx_session

from business_manager.import_data.services import CuteViewSet, CutePageNumberPagination
from business_manager.collection.services import collection_extra_data
from business_manager.collection.general_views import apply_trans_status
from business_manager.collection.serializers import (
    CollectionRecordSerializer, CollectionRecordTagSerializer, QualitySerializer
)
from business_manager.util.import_file_util import ImportFile
from business_manager.util.tkdate import get_yestoday,get_tomorrow,get_first_day_of_month,get_today,get_first_day_of_week

employee_groups_name = u'催收主管,催收M1,催收M2,催收M3,催收M4,催收M5,催收M5以上'.split(',')

employee_groups = {
    "a": "", 
    "b": "催收M1",
    "c": "催收M2",
    "d": "催收M3",
    "e": "催收M4",
    "g": "催收M5",
    "h": u"催收M5以上"
}

coll_level_dict = {
    'a': 'all', 
    'b': 'm1', 
    'c': 'm2', 
    'd': 'm3', 
    'e': 'm4', 
    'g': 'm5', 
    'h': 'm5+'
}

phone_pattern = re.compile('1\d{10}')

@require_http_methods(['POST'])
@csrf_exempt
def report_log(request):
    """"""
    message = request.POST.get('raw_data', '')
    Log().info(u'通话日志：%s' % message)
    return JsonResponse({
        'code': 0,
        'msg': u'日志记录成功'
    })


def add_message_record(request, type=CollectionRecord.MESSAGE,data=None):
    """发送短信后添加催记"""
    if not data:
        return 
    try:
        employee = Employee.objects.get(user=request.user)
        CollectionRecord.objects.create(
            record_type=type,
            object_type=data.get('phone', ''),
            collection_note=data.get('content', ''),
            create_by=employee,
            apply=data.get('order', None)
            )
        return True
    except Exception, e:
        print(e)
        return 


@require_http_methods(['POST'])
@csrf_exempt
def send_message(request):
    '''发送短信接口'''
    phone = request.POST.get('phone', '').strip()
    content = request.POST.get('content', '').strip()
    message_template_id = request.POST.get('message_template_id', '').strip()
    if len(content) > 256:
        return JsonResponse({
            'code': -1,
            'msg': u'内容超过了256个字符'
        })
    print request.POST.get('apply_id')
    print request.POST
    order = Apply.objects.filter(pk=request.POST.get('apply_id')).first()

    message_template = None
    if message_template_id:
        message_template = Messagetemplate.objects.get(id=message_template_id)

    employee = Employee.objects.get(user=request.user)
    collection_record_data = dict(
        record_type=CollectionRecord.MESSAGE,
        object_type=phone,
        create_by=employee,
        # collection_note=content,
        promised_repay_time=None,
        apply=order,
        message_template=message_template,
    )

    extra_data = collection_extra_data(order)
    collection_record_data.update(**extra_data)
    print collection_record_data


    try:
        client = MessageClient(settings.MESSAGE_SERVER['HOST'], settings.MESSAGE_SERVER['PORT'], logger=None)
        res = client.dd_send_message(phone, content.encode('utf-8'))
        print 'send result: ', res
        if res:
            code = 0
            msg = u'短信发送成功'
        else:
            content = u'短信: 发送失败 --  '  + content
            code = -1
            msg = u'短信发送失败'

        collection_record_data['collection_note'] = content
        record = CollectionRecord(**collection_record_data)
        record.save()
        if res:
            apply_trans_status(order, record)
            if 's2' != order.status:
                order.status = 's2'


        return JsonResponse({
            'code': code,
            'msg': msg
        })
    except Exception, e:
        # add_message_record({"phone": phone, "content": content, "order": order, "user": request.user, "status": False})
        print(e)
        return JsonResponse({
            'code': -1,
            'msg': u'短息发送失败'
        })


@require_http_methods(['GET'])
def get_msg_template(request):
    '''拉取短信模板'''
    msg_id = request.GET.get('msg_id')
    if msg_id:
        msg = get_object_or_404(Messagetemplate, pk=msg_id)
        data = {
            'id': msg.id,
            'content': msg.content,
            'title': msg.title,
        }
        return JsonResponse({
            'code': 0,
            'msg': u'返回模板成功',
            'data': data
        })
    
    messages = Messagetemplate.objects.all()
    data = list()
    for message in messages:
        data.append({
            'id': message.id,
            'title': message.title,
            'content': message.content
        })
    return JsonResponse({
        'code': 0,
        'msg': u'返回模板成功',
        'data': data
    })


@require_http_methods(['GET'])
def get_collection_type(request):
    """"""
    return JsonResponse({
        "code": 0,
        "msg": "",
        "collection_type": {
            "a": "所有",
            "b": "M1",
            "c": "M2",
            "d": "M3",
            "e": "M4",
            "g": "M5",
            "h": "M5+"
        }
    })


@require_http_methods(['GET'])
def get_lending_channels(request):
    """贷款渠道"""
    users = User.objects.values('channel').distinct()
    results = list()
    for u in users:
        if u.get('channel'):
            results.append(u.get('channel', ''))

    return JsonResponse({
        "code": 0,
        "msg": "",
        "channels": results
    })


@require_http_methods('GET')
def get_collection_personnel(request):
    """"""
    # 通过用户组过滤employee
    group_type = request.GET.get('type')
    platform = get_employee_platform(request)[0].name
    if group_type and u'a' != group_type:
        ges = list()
        if group_type.strip() in employee_groups.keys():
            group = EmployeeGroup.objects.filter(group_name=employee_groups[group_type], platform = platform).first()
            if group:
                employees = group.employee_set.all()
                for employee in employees:
                    ges.append({
                        "id": employee.id, "name": employee.username, "username": employee.user.username, 'is_active':employee.user.is_active, 'coll_level': coll_level_dict.get(group_type, '')})
    else:
        permission_employees = list()
        for group_name in employee_groups_name:
            p = re.search(r'M(\d{1})(.*) | S(\d{1})(.*)', group_name)
            coll_level = group_name
            if p:
                if p.group(1) and p.group(2):
                    coll_level = 'm5+'
                elif p.group(1) and not p.group(2):
                    coll_level = 'm' + p.group(1)
            employees = Employee.objects.filter(group_list__group_name=group_name, platform_list__in = get_employee_platform(request))
            for e in employees:
                permission_employees.append({"id": e.id, "name": e.username, 'username': e.user.username, 'is_active':e.user.is_active, 'coll_level': coll_level})
        ges = permission_employees
    return JsonResponse({
        "code": 0,
        "msg": "",
        "employees": ges
    })


@require_http_methods(['GET'])
def get_user_info(request):
    """"""
    order = get_object_or_404(Apply, pk=request.GET.get('apply_id'))
    card = order.create_by.bankcard_set.first()

    banks = {
        "CCB": u"中国建设银行",
        "ABC": u"中国农业银行",
        "ICBC": u"中国工商银行", 
        "BOC": u"中国银行",
        "CMBC": u"中国民生银行",
        "CMB": u"招商银行",
        "CIB": u"兴业银行",
        "BCM": u"交通银行",
        "CEB": u"中国光大银行",
        "GDB": u"广东发展银行",
        "CITIC": u"中信银行",
        "SPDB": u"浦发银行",
        "HXB": u"华夏银行",
        "PAB": u"平安银行",
    }
    bank_name = ''
    if card:
        bank_name = card.bank_name or ''
        if card.bank_name in banks.keys():
            bank_name = banks.get(card.bank_name)
    info = {
        "name": order.create_by.name if order.create_by else '',
        "phone": order.create_by.phone_no if order.create_by else '',
        "id_no": order.create_by.id_no if order.create_by else '',
        "payment": order.repayment.apply_amount if order.repayment else order.amount,
        "deduction_amount": order.repayment.repay_amount if order.repayment else order.rest_repay_money,
        "loan_type": order.repayment.strategy.name if order.repayment else u'12个月',
        "bank_name": bank_name,
        "bank_number": card.card_number if card else '',
        "employee_name": order.employee.username if order.employee else '',
        "employee_id": order.employee.id if order.employee else '',
        "collection_level": order.type
    }
    return JsonResponse({
        "code": 0,
        "msg": u"返回用户信息成功",
        "user_info": info
    })


class CollectionDataProvider(object):
    """"""
    ORDERS = {
        "created_at": 'create_at',
        'payment': 'rest_repay_money', 
        'promised_payment_time': 'last_commit_at',
        'overdu_days': 'overdue_days',
        'loan_amount': 'repayment__apply_amount',
        'collection_log_time': 'last_commit_at'
    }

    def object_filter(self, request=None, owner=None):
        """过滤"""
        # filter by overdu type
        overdu_type = request.GET.get('type', '').strip()
        query_type = Q()
        if overdu_type in ['b', 'c', 'd', 'e', 'g', 'h']:
            query_type = Q(type=overdu_type)
        # if 'm1' == overdu_type:
        #     query_type = Q(type='b')
        # elif 'm2' == overdu_type:
        #     query_type = Q(type='c')
        # elif 'm3' == overdu_type:
        #     query_type = Q(type='d')
        # elif 'm4' == overdu_type:
        #     query_type = Q(type='e')
        # elif 'm5' == overdu_type:
        #     query_type = Q(type='g')
        # elif 'm6' == overdu_type:
        #     query_type = Q(type='h')
        else:
            query_type = Q(type__in=['a', 'b', 'c', 'd', 'e', 'g', 'h'])

        # filter by status.
        status = request.GET.get('status', '').strip()
        query_status = Q()
        if 'wait_distribution' == status:
            query_status = Q(status='s1') | Q(status='0')
        elif 'not_accept' == status:
            query_status = Q(status='s2') | Q(status='i')
        elif 'processing' == status:
            query_status = Q(status='s3') | Q(status='ci')
        elif 'wait_check' == status:
            query_status = Q(status='k')
        elif 'check_failed' == status:
            query_status = Q(status='t')
        elif 'collection_success' == status:
            query_status = Q(status='8') | Q(status='9')
        elif 'repay_failed' == status:
            query_status = Q(status='c')
        elif 'partial_success' == status:
            query_status = Q(status='d')
        elif 'repay_error' == status:
            query_status = Q(status='o')
        elif 'lost_contact' == status:
            query_status = Q(status=Apply.LOST_CONTACT)
        elif 'recall_fail' == status:
            query_status = Q(status=Apply.RECALL_FAIL)
        elif 'recall_success' == status:
            query_status = Q(status=Apply.RECALL_SUCCESS)
        elif 'repair_fail' == status:
            query_status = Q(status=Apply.REPAIR_FAIL)
        elif 'renew' == status:
            query_status = Q(status=Apply.RENEW)
        else:
            pass
        
        # filter by channel.
        lending_channel = request.GET.get('lending_channels', '').strip()
        query_channel = Q()
        if lending_channel:
            query_channel = Q(create_by__channel=lending_channel)
        
        # filter by Employee.
        collection_employee = request.GET.get('collection_employee', 0)
        query_employee = Q()
        if collection_employee and u'' != collection_employee:
            employee = Employee.objects.filter(pk=collection_employee).first()
            if employee:
                query_employee = Q(employee=employee)
        
        query_owner = Q()
        if owner:
            try:
                _owner = Employee.objects.get(user_id=request.user.id)
                print 'owner: ', _owner
                query_owner = Q(employee=_owner)
            except Exception, e:
                print e

        # filter by time
        time = request.GET.get('time')
        start_time = end_time = None
        query_time = Q()
        if 'today' == time:
            start_time = get_today()
            end_time = get_tomorrow()
        elif 'yestoday' == time:
            start_time = get_yestoday()
            end_time = get_today()
        elif 'toweek' == time:
            start_time = get_first_day_of_week()
            end_time = get_tomorrow()
        elif 'tomonth' == time:
            start_time = get_first_day_of_month()
            end_time = get_tomorrow()
        elif 'other' == time:
            start_time = request.GET.get('start_time')
            end_time = request.GET.get('end_time')
        
        if start_time and end_time:
            query_time = Q(create_at__lte=end_time, create_at__gte=start_time)
        

        if 'collection_success' == status and start_time and end_time:
            query_time = Q(real_repay_time__lte=end_time, real_repay_time__gte=start_time)
        elif 'partial_success' == status and start_time and end_time:
            query_time = Q(real_repay_time__lte=end_time, real_repay_time__gte=start_time)

        payment_info = request.GET.get('payment_range')
        print '------', payment_info
        if payment_info:
            payment_min, payment_max = payment_info.split(',')
            try:
                payment_min = int(payment_min)
                payment_max = int(payment_max)
            except:
                payment_min = -1
                payment_max = -1
            payment_start = Q(rest_repay_money__gte=payment_min) if payment_min > 0 else Q()
            payment_end = Q(rest_repay_money__lte=payment_max) if payment_max > 0 else Q()
            query_payment = payment_start & payment_end
        else:
            query_payment = Q()
            
        # filter by overdu_days
        overdu_days = request.GET.get('overdu_days')
        query_overdu_days = Q()
        if overdu_days:
            overdu_days = json.loads(overdu_days)
            start_day = overdu_days.get('start_day')
            end_day = overdu_days.get('end_day')
            if start_day and end_day and end_day >= start_day >= 0:
                query_overdu_days = Q(overdue_days__gte=start_day, overdue_days__lte=end_day)
        
        query_platform = Q()
        user_platform = get_employee_platform(request)
        print 'user_platform'
        print user_platform



        if len(user_platform) > 1:
            platform_param = request.GET.get('platform', '')
            if platform_param:
                query_platform = Q(platform = platform_param)
            else:
                query_platform = Q(platform__in = user_platform.values_list('name', flat = True))
        else:
            query_platform = Q(platform = user_platform[0].name)

        print query_platform
        query_product = Q()
        product = request.GET.get('product', '')
        if product:
            query_product = Q(product = product)
        applies = Apply.objects.filter(query_owner, query_type, query_status, query_channel, query_payment,
                                       query_employee, query_time, query_overdu_days, query_platform, query_product)
        
        search = request.GET.get('search', '').strip()
        if search:
            applies = applies.filter(Q(create_by__name__contains=search) | Q(create_by__phone_no=search) | Q(repayment__order_number=search) | Q(create_by__id_no=search))
            if phone_pattern.match(search):
                contact_list = ContactInfo.objects.filter(Q(phone_no=search))
                tmp_user_list = []
                for contact in contact_list:
                    if contact.owner not in tmp_user_list:
                        tmp_user_list.append(contact.owner)
                        tmp_apply = Apply.objects.filter(Q(create_by=contact.owner), query_type, query_status, query_channel, query_employee, query_time, query_overdu_days)
                        applies |= tmp_apply
        # print query_owner, query_type, query_status, query_channel, query_employee, query_time, query_overdu_days
        # print applies
        return applies
    
    def sort(self, request=None, queryset=None):
        """排序"""
        order_name, order_way = request.GET.get('order_name', '').strip(), request.GET.get('order_way', '').strip()
        if order_name and order_way:
            if 'asc' == order_way:
                return queryset.order_by(CollectionDataProvider.ORDERS.get(order_name))
            elif 'desc' == order_way:
                return queryset.order_by('-' + CollectionDataProvider.ORDERS.get(order_name))
        return queryset.order_by('-id')

        
    def pagination(self, request=None, queryset=None):
        """分页"""
        page, page_size = request.GET.get('page'), request.GET.get('page_size')
    
        if page and page_size:
            start = int(page) * int(page_size)
            return queryset[start-int(page_size):start]
        return queryset[:15]

    def genrate_result(self, request=None, owner=False):
        """"""
        _applies = self.object_filter(request, owner=owner)
        print _applies
        sorted_applies = self.sort(request, _applies)
        print 'sort'
        applies = self.pagination(request, sorted_applies)
        print 'pagination'
        # url = request.build_absolute_uri()
        url = settings.DOMAIN + request.get_full_path()
        url = url.replace('all_collection_json', 'down_tables').replace('my_collection_json', 'down_tables')
        order_count = _applies.count()
        # total_amount = _applies.aggregate(Sum('repayment__apply_amount')).get('repayment__apply_amount__sum')
        total_amount = _applies.aggregate(Sum('rest_repay_money')).get('rest_repay_money__sum')
        data = list()
        print total_amount
        print applies.count()
        # _apply = Apply.objects.filter(id=228).first()
        # print _apply.employee

        for al in applies:
            try:
                coll_records = al.collectionrecord_set.all().order_by('-id')
                log_time = ''
                if coll_records:
                    log_time = coll_records[0].create_at.strftime("%Y-%m-%d")
                status_name = None
                payment = 0
                if al.status in ['8', '9', 's4']:
                    status_name = u'催收完成'
                    # payment = al.real_repay_money
                elif al.status in ['s1', '0']:
                    status_name = u'待分配'
                elif al.status in ['s3', 'ci']:
                    status_name = u'已受理'
                elif al.status in ['s2', 'i']:
                    status_name = u'未受理'

                # 续期状态下, 需要使用到的 倒计时天数.
                renew_days = 0
                renew_installment = InstallmentDetailInfo.objects.filter(repayment=al.repayment).first()
                if renew_installment:
                    renew_days = (renew_installment.should_repay_time - datetime.now()).days

                data.append({
                    'id': al.id,
                    'order_number': al.repayment.order_number if al.repayment else 0,
                    'name': al.create_by.name if al.create_by else '',
                    'created_at': al.create_at.strftime("%Y-%m-%d") if al.create_at else '',
                    'payment': payment if payment else al.rest_repay_money,
                    'promised_payment_time': al.promised_repay_time.strftime("%Y-%m-%d") if al.promised_repay_time else '',
                    'overdu_days': al.overdue_days,
                    'loan_amount': al.repayment.apply_amount if al.repayment else 0,
                    'channel': al.create_by.channel if al.create_by else '',
                    'dispathch_name': al.employee.username if al.employee else '',
                    'collection_log_time': al.collection_record_time.strftime("%Y-%m-%d") if al.collection_record_time else log_time,
                    'status': status_name if status_name else al.get_status_display(),
                    'url': '/collection/info/{}'.format(al.id),
                    'renew_days': renew_days,
                    'info_repair': 1 if al.extra_info == Apply.INFO_REPAIR else 0,
                })
            except Exception, e:
                print e
                continue
        return {'url': url, 'order_count': order_count, 'total_amount': total_amount, 'data': data}


@require_http_methods('GET')
def all_collection(request):
    """"""
    results = CollectionDataProvider().genrate_result(request)
    if isinstance(results, dict):
        results.update({'code': 0, 'msg': ''})
        return JsonResponse(results)
    return JsonResponse({
        'code': 0,
        'msg': '', 
        'data': []
    })


@require_http_methods(['GET'])
def my_collection(request):
    ''''''
    results = CollectionDataProvider().genrate_result(request, owner=True)
    if isinstance(results, dict):
        results.update({'code': 0, 'msg': ''})
        return JsonResponse(results)
    return JsonResponse({
        'code': 0,
        'msg': '',
        'data': []
    })


@require_http_methods(['POST'])
@csrf_exempt
@page_permission(check_employee)
def order_allocation(request):
    '''分配催收订单'''
    order_id, employee_id = request.POST.get('apply_id'), request.POST.get('employee_id')
    print order_id, employee_id
    try:
        order = Apply.objects.get(pk=order_id)
        employee = Employee.objects.get(pk=employee_id)
        if order and employee:
            if order.status in (Apply.LOST_CONTACT, Apply.REPAIR_FAIL, Apply.RECALL_FAIL, Apply.RECALL_SUCCESS):
                return JsonResponse({
                     "code": -1,
                     "msg": u"订单不属于待分配状态"
                })
            # if order.status not in ['0', 's1', 's2', 's3', 'i']:
            #     return JsonResponse({
            #         "code": -1,
            #         "msg": u"订单不属于待分配状态"
            #     })
            if order.employee:
                report_collection(apply=order, status_name=u'失效')
            if order.status in ['0', 's1', 's2', 's3', 'i']:
                order.status = 's2'
            order.employee = employee
            order.save()

            try:
                create_by = Employee.objects.filter(user_id=request.user.id).first()
                collection_record_data = dict(
                    record_type=CollectionRecord.DISPATCH,
                    object_type=CollectionRecord.OTHER,
                    collection_note=u"分配订单给{}".format(employee.username),
                    create_by=create_by,
                    apply=order
                )

                extra_data = collection_extra_data(order)
                collection_record_data.update(**extra_data)
                print collection_record_data

                CollectionRecord.objects.create(**collection_record_data)
                report_collection(order)

            except Exception, e:
                print u'分配订单：\n', e
            return JsonResponse({
                'code': 0,
                'msg': u'分配催收成功'
            })
        return JsonResponse({
            "code": -1,
            "msg": u'订单或员工不存在'
        })
    except Exception, e:
        print e
        return JsonResponse({
            'code': -1,
            'msg': u'订单或员工不存在',
        })


@require_http_methods(['GET'])
def pull_repay_type(request):
    """返回还款类型列表"""
    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": [{'key1':'installment', 'key2':u'期款'}, ]
    })


@require_http_methods(['GET'])
def pull_collection_record(request):
    '''拉取给定订单下所有催记'''
    order_id = request.GET.get('apply_id')
    type = request.GET.get('type', -1)
    order_way = request.GET.get('order_way', '').strip()
    page, page_size = request.GET.get('page'), request.GET.get('page_size')
    print 'in pull_collection_record'

    query_type = Q()
    type = str(type)
    if type in ['0', '1', '2', '3', '4', '5', '6', '7']:
        query_type = Q(record_type=type)

    apply = get_object_or_404(Apply, id=order_id)
    apply_ids = Apply.objects.filter(repayment=apply.repayment)

    if 'asc' == order_way:
        records = CollectionRecord.objects.filter(apply__in=apply_ids).filter(query_type).order_by('create_at')
    else:
        records = CollectionRecord.objects.filter(apply__in=apply_ids).filter(query_type).order_by('-create_at')

    print '111111'
    _count = records.count()
    if page and page_size:
        start = int(page) * int(page_size)
        records = records[start-int(page_size):start]
    else:
        records = records[:5]

    data = list()
    for record in records:
        image_url = ''
        if record.check_apply:
            print record.id
            image_url = record.check_apply.pic

        print record.id
        print record
        print record.create_by
        data.append({
            'id': record.id,
            'record_type': record.get_record_type_display(),
            'collector': record.create_by.username,
            'add_time': record.create_at.strftime('%Y-%m-%d %H:%M:%S') if record.create_at else '',
            'promised_repay_time': record.promised_repay_time.strftime('%Y-%m-%d %H:%M:%S') if record.promised_repay_time else '',
            'notes': record.collection_note,
            'image_url': image_url,
        })
    return JsonResponse({
        'code': 0,
        'msg': u'拉取催记成功',
        'data': data,
        'count': _count
    })


@require_http_methods(['GET'])
def get_relation_info(request):
    """"""
    _id = request.GET.get('apply_id')
    if _id:
        order = get_object_or_404(Apply, pk=request.GET.get('apply_id'))
        users = ContactInfo.objects.filter(owner=order.create_by)
        if order.type in [Apply.COLLECTION_M0, Apply.COLLECTION_M1]:
            users = users[:100]

        data = list()
        data.append({
            'name': order.create_by.name, 
            'phone': order.create_by.phone_no, 
            'relationship': u'本人', 
            'relationshipid': None,
            'address': '',
            'id_no': order.create_by.id_no,
            'relationship_desc': ''
            })
        for user in users:
            relationship = user.get_relationship_display()
            if user.relationship == 0:
                relationship = user.relationship_desc or relationship

            data.append({
                'name': user.name,
                'phone': user.phone_no,
                'relationship': relationship or u'其他',
                'relationshipid': user.relationship,
                'address': user.address,
                'id_no': user.id_no,
                'relationship_desc': user.relationship_desc
            })

        return JsonResponse({
            'code': 0,
            'msg': u'返回成功',
            'relation_info': data
        })
    return JsonResponse({
        'code': -1,
        'msg': u'参数错误'
    })


@require_http_methods(['POST'])
@csrf_exempt
def add_or_update_relationship(request):
    '''更新或添加联系人'''
    order = get_object_or_404(Apply, pk=request.POST.get('apply_id'))
    
    # update.
    old_phone = request.POST.get('old_phone').strip()
    if old_phone:
        info = ContactInfo.objects.filter(owner=order.create_by, phone_no=old_phone).first()
        if info: 
            info.name = request.POST.get('name', info.name)
            info.address = request.POST.get('address', info.address)
            info.id_no = request.POST.get('id_no', info.id_no)
            info.relationship = request.POST.get('relationship', info.relationship)
            info.relationship_desc = request.POST.get('relationship_desc', info.relationship_desc)
            info.phone_no = request.POST.get('phone', info.phone_no)
            info.save()
            return JsonResponse({
                'code': 0,
                'msg': u'更新信息成功'
            })
        return JsonResponse({
            'code': -1,
            'msg': u'订单号或电话号码错误'
        })
    
    # create.
    try:
        _phone = request.POST.get('phone', '')
        info = ContactInfo.objects.filter(phone_no=_phone).first()
        if info:
            return JsonResponse({
                'code': -1,
                'msg': u'该号码已经存在'
            })
        ContactInfo.objects.create(
            name=request.POST.get('name', '').strip(),
            address=request.POST.get('address', '').strip(),
            id_no=request.POST.get('id_no', '').strip(),
            phone_no=request.POST.get('phone', '').strip(),
            relationship=request.POST.get('relationship', 0),
            relationship_desc=request.POST.get('relationship_desc', '').strip(),
            owner=order.create_by
        )
        return JsonResponse({
            'code': 0,
            'msg': u'创建成功'
        })
    except Exception, e:
        print(u'创建关系联系人信息失败：\n', e)
        return JsonResponse({
            'code': -1,
            'msg': u'创建联系人失败'
        })


def get_strategy_type(strategy):
    if strategy.installment_type == 1:
        name = u'%s%s' % (strategy.installment_count, u'月')
    elif strategy.installment_type == 2:
        name = u'%s%s' % (strategy.installment_days, u'天')

    return name

@require_http_methods(['GET'])
def pull_loan_info(request):
    """拉取贷款信息"""
    _apply = get_object_or_404(Apply, pk=request.GET.get('apply_id'))
    
    if _apply.repayment:
        repayment = _apply.repayment
        _repay = {
            'order_number': repayment.order_number,
            'apply_amount': repayment.apply_amount,
            'exact_amount': repayment.exact_amount,
            'rest_amount': repayment.rest_amount,
            'repay_status': repayment.repay_status,
            'apply_time': repayment.apply_time.strftime('%Y-%m-%d %H:%M:%S') if repayment.apply_time else ''
        }

        install_list = InstallmentDetailInfo.objects.filter(repayment=repayment).order_by('installment_number')
        installment_records = InstallmentRecord.objects.filter(repayment=repayment).order_by('installment_number')

        installs = list()
        for install in install_list:
            ins_repay_amount = install.real_time_should_repay_amount or install.should_repay_amount + install.repay_overdue
            installs.append({
                "installment_number": install.installment_number,
                "should_repay_time": install.should_repay_time.strftime('%Y-%m-%d') if install.should_repay_time else '',
                "real_repay_time": install.real_repay_time.strftime('%Y-%m-%d') if install.real_repay_time else '',
                "should_repay_amount": ins_repay_amount,
                "real_repay_amount": install.real_repay_amount,
                "repay_status": install.repay_status,
                # "repay_status": install.get_repay_status_display(), 
                "overdue_days": install.overdue_days,
                # "overdue_status": install.repay_type if install.repay_status else 0,
                "overdue_status": employee_groups.get(gct(install.overdue_days), ''),
                # "stratety_type": u'{}个月'.format(repayment.installment_count),
                "stratety_type": get_strategy_type(repayment.strategy),
                "overdue_repay": install.repay_overdue,
                # "should_repay": ins_repay_amount - install.real_repay_amount,
                "should_repay": ins_repay_amount - install.real_repay_amount - install.reduction_amount,
                "review_staff_name": get_review_staff_name(install),
                "reduction_amount": install.reduction_amount,
            })

        for install in installment_records:
            ins_repay_amount = install.real_time_should_repay_amount or install.should_repay_amount + install.repay_overdue
            installs.append({
                "installment_number": install.installment_number,
                "should_repay_time": install.should_repay_time.strftime('%Y-%m-%d') if install.should_repay_time else '',
                "real_repay_time": install.real_repay_time.strftime('%Y-%m-%d') if install.real_repay_time else '',
                "should_repay_amount": ins_repay_amount,
                "real_repay_amount": install.real_repay_amount,
                "repay_status": install.repay_status,
                # "repay_status": install.get_repay_status_display(), 
                "overdue_days": install.overdue_days,
                # "overdue_status": install.repay_type if install.repay_status else 0,
                "overdue_status": employee_groups.get(gct(install.overdue_days), ''),
                # "stratety_type": u'{}个月'.format(repayment.installment_count),
                "stratety_type": get_strategy_type(install.strategy),
                "overdue_repay": install.repay_overdue,
                "should_repay": ins_repay_amount - install.real_repay_amount - install.reduction_amount,
                "review_staff_name": get_review_staff_name(install),
                "reduction_amount": install.reduction_amount,
            })


        # print installs
        _repay.update({
            "install_list": installs,
            "count": len(installs)
            })

        return JsonResponse({
            'code': 0,
            'msg': u'返回成功',
            'data': _repay
        })
    return JsonResponse({
        'code': 0,
        'msg': u'该订单下无贷款信息',
        'data': []
    })


@require_http_methods(['GET'])
def pull_collection_info(request, apply_id=None):
    '''拉取催收信息(用户基本信息)'''
    order = get_object_or_404(Apply, pk=apply_id)

    # try:
    if 1:
        if order.create_by.submit_modules_state:
            key = order.create_by.id_no
            url = 'http://{0}:{1}/get_submitted_profile?search_key={2}&product={3}&platform={4}'.format(
                settings.WEB_HTTP_SERVER['HOST'],
                settings.WEB_HTTP_SERVER['PORT'],
                key,
                order.product,
                order.platform
            )
            print url
            data = requests.get(url, timeout=3)
            print data.text
            data = data.json()
            # .json()
        else:
            url = 'http://{0}:{1}/get_import_user_info?uin={2}'.format(
                settings.WEB_HTTP_SERVER['HOST'],
                settings.WEB_HTTP_SERVER['PORT'],
                order.create_by.id
            )
            # print url
            data = requests.get(url, timeout=3)
            # print data.text
            data = data.json()
        _code = data.get('code', 0)
        msg = None
        if _code:
            msg = u'获取用户信息失败, {}'.format(data.get('msg', ''))
        p = re.search(r'(m\d{1}.?)', order.get_type_display())
        if p:
            data.update({"collection_type": p.group(1)})
        else:
            data.update({"collection_type": ''})
        return JsonResponse({
            'code': _code if _code else 0,
            'msg': msg if msg else u'获取用户信息成功',
            'data': data
        })
    # except Exception, e:
        # print(u'获取用户基本信息失败：\n', e)
        # return JsonResponse({
            # 'code': -1,
            # 'msg': u'获取用户信息失败: {}'.format(e)
        # })


# 弃用
@require_http_methods(['POST'])
@csrf_exempt
@page_permission(check_employee)
def add_record(request):
    """添加催记"""

    employee = get_object_or_404(Employee, user_id=request.user.id)
    order = get_object_or_404(Apply, pk=request.POST.get('apply'))
    if order.employee != employee:
        return JsonResponse({
            "code": -1,
            "msg": u"您没有为该订单添加催记权限，请先分配"
        })
    
    will_repay_time = request.POST.get('time')

    try:
        if will_repay_time:
            promised_repay_time = datetime.strptime(will_repay_time, "%Y-%m-%d %H:%M")
        else:
            promised_repay_time = None
        print '--------', promised_repay_time
        order.promised_repay_time = promised_repay_time
        order.collection_record_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if order.status in ['s2', 'i']:
            order.status = 's3'
        order.save()

        _records = CollectionRecord.objects.filter(apply=order, create_by=employee, record_type=CollectionRecord.COLLECTION)
        if not _records:
            try:
                report_collection(apply=order)
            except:
                pass    

        record = CollectionRecord(
            record_type=CollectionRecord.COLLECTION,
            object_type=request.POST.get('object'),
            create_by=employee,
            collection_note=request.POST.get('content', '').strip(),
            promised_repay_time=promised_repay_time, 
            apply=order
            )
        print record.promised_repay_time
        record.collection_note = record.get_object_type_display() + '--' + request.POST.get('content', '').strip()
        record.save()
        try:
            print '*' * 32
            report_collection(apply=order, status_name=u'催记')
        except:
            pass
        return JsonResponse({
            'code': 0,
            'msg': u'添加催记成功'
        })
    except Exception, e:
        print u'添加催记失败：\n', e
        return JsonResponse({
            'code': -1,
            'msg': u'添加催记失败'
        })

@require_http_methods(['POST'])
def wx_add_record(request):
    """微信端添加催记"""

    session = request.POST.get('session')
    employee_id = check_wx_session(session)
    if not employee_id:
        return JsonResponse({
            "code": -1,
            "msg": u"请先登录"
        })
    employee = get_object_or_404(Employee, pk = employee_id)
    order = get_object_or_404(Apply, pk=request.POST.get('apply'))
    if order.employee != employee:
        return JsonResponse({
            "code": -1,
            "msg": u"您没有为该订单添加催记权限，请先分配"
        })
    
    will_repay_time = request.POST.get('time')

    try:
        if will_repay_time:
            promised_repay_time = datetime.strptime(will_repay_time, "%Y-%m-%d %H:%M")
        else:
            promised_repay_time = None
        order.last_commit_at = promised_repay_time
        if order.status in ['s2', 'i']:
            order.status = 's3'
        order.save()

        _records = CollectionRecord.objects.filter(apply=order, create_by=employee, record_type=CollectionRecord.COLLECTION)
        if not _records:
            try: 
                report_collection(apply=order)
            except:
                pass    

        record = CollectionRecord(
            record_type=CollectionRecord.COLLECTION,
            object_type=request.POST.get('object'),
            create_by=employee,
            collection_note=request.POST.get('content', '').strip(),
            promised_repay_time=promised_repay_time, 
            apply=order
            )
        record.collection_note = record.get_object_type_display() + '--' + request.POST.get('content', '').strip()
        record.save()
        try:
            report_collection(order, status_name=u'催记')
        except:
            pass
        return JsonResponse({
            'code': 0,
            'msg': u'添加催记成功'
        })
    except Exception, e:
        print u'添加催记失败：\n', e
        return JsonResponse({
            'code': -1,
            'msg': u'添加催记失败'
        })




@require_http_methods(['GET'])
def down_tables(request):
    """下载报表(需要删除之前的文件)"""
    
    if not settings.ALLOW_DOWNLOAD_REPORT:
        return JsonResponse({
            "code": -1,
            "msg": u'报表下载功能暂停使用'
        })
    try:
        os.system('rm orders*')
    except:
        pass
    applies = CollectionDataProvider().object_filter(request)

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S').replace('-', '').replace(' ', '').replace(':', '')
    path = 'orders' + now + '.xlsx'
    wb = Workbook()
    ws = wb.create_sheet('orders', 0)

    ws.title = u'导出催收订单列表'
    ws.append([u'ID', u'工单号', u'用户名',  u'进件日期', u'期款', u'承诺还款时间', u'逾期天数', 
               u'贷款金额', u'贷款渠道', u'处理人', u'催记时间', u'处理状态'])

    for al in applies:
        try:
            coll_records = al.collectionrecord_set.all().order_by('-id')
            ws.append([al.id, al.repayment.order_number if al.repayment else '', al.create_by.name if al.create_by else '', al.create_at.strftime('%Y-%m-%d'), al.rest_repay_money / 100.0, al.promised_repay_time.strftime('%Y-%m-%d') if al.promised_repay_time else '', al.overdue_days, al.repayment.apply_amount / 100.0 if al.repayment else 0, al.create_by.channel if al.create_by else '', al.employee.username if al.employee else u'', coll_records[0].create_at.strftime('%Y-%m-%d') if coll_records else '', al.get_status_display()])
        except Exception, e:
            print al
            print(u'写入xlsx：\n', e)
            continue
    total_amount = applies.aggregate(Sum('repayment__apply_amount')).get('repayment__apply_amount__sum') / 100.0
    total_rest_amount = applies.aggregate(Sum('rest_repay_money')).get('rest_repay_money__sum') / 100.0
    # total_amount = sum([a.repayment.apply_amount for a in applies if a.repayment])
    ws.append([u'总计', '', '', '', total_rest_amount, '', '', total_amount, '', '', '', ''])
    wb.save(path)
    
    response = StreamingHttpResponse(FileWrapper(open(path), 8192), content_type='application/vnd.ms-excel')
    response['Content-Length'] = os.path.getsize(path)
    response['Content-Disposition'] = 'attachment; filename={}'.format(path)
    return response


def test_down_tables(request):
    from django.db import connection
    from openpyxl.writer.excel import save_virtual_workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet('orders', 0)
    ws.append([u'ID', u'用户名', u'工单号',  u'进件日期', u'期款', u'承诺还款时间', u'逾期天数',
               u'贷款金额', u'贷款渠道', u'处理状态', u'处理人', u'催记时间'])

    where = get_where(request)
    query = (
        "select a.id, b.name, c.order_number, a.create_at, "
        "a.real_repay_money, a.promised_repay_time, a.overdue_days, "
        "c.apply_amount, b.channel, a.status, d.username, max(e.create_at) record_time "
        "from apply a left join `user` b on a.create_by_id = b.id "
        "left join repaymentinfo c on a.repayment_id = c.id "
        "left join employee d on a.employee_id = d.id "
        "left join review_collectionrecord e on e.apply_id = a.id "
        "where %s GROUP BY a.id" % where
    )
    print query
    status_map = {
        '8': u'催收完成', 8: u'催收完成',
        '9': u'催收完成', 9: u'催收完成',
        's4': u'催收完成',
        's1': u'待分配', '0': u'待分配',
        0: u'待分配',
        's3': u'已受理', 'ci': u'已受理',
        's2': u'未受理', 'i': u'未受理',
    }
    with connection.cursor() as cursor:
        cursor.execute(query)
        # map(lambda item: ws.append(item), cursor)
        for item in cursor:
            item = list(item)
            if item[4]:
                item[4] /= 100.0
            else:
                item[4] = 0
            if item[7]:
                item[7] /= 100.0
            else:
                item[7] = 0
            if item[9]:
                item[9] = status_map.get(item[9])
            ws.append(item)
    rsp = StreamingHttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    rsp['Content-Disposition'] = 'attachment; filename=order.xlsx'
    return rsp


def get_where(request):
    """ a: apply, b: user, c: repaymentinfo, d: employee, e: collectionrecord
    """
    where = []
    params = request.GET.dict()
    overdue_type = params.get("type", "").strip()
    if overdue_type in ['b', 'c', 'd', 'e', 'g', 'h']:
        type_condition = "a.type = %s" % overdue_type
    else:
        type_condition = "a.type in %s" % str(('a', 'b', 'c', 'd', 'e', 'g', 'h'))
    where.append(type_condition)

    status_map = {
        "wait_distribution": "('s1', 0)",
        "not_accept": "('s2', 'i')",
        "processing": "('s3', 'ci')",
        "wait_check": "('k')",
        "check_failed": "('t')",
        "collection_success": "(8, 9)",
        "repay_failed": "('c')",
        "partial_success": "('d')",
        "repay_error": "('o')"
    }
    status = params.get("status", "").strip()
    if status:
        where.append("a.status in %s" % status_map.get(status))

    channel = params.get("lending_channels", "").strip()
    if channel:
        where.append("b.channel=%s" % channel)

    employee = params.get('collection_employee', 0)
    if employee and employee != '0':
        where.append("a.employee_id=%s" % employee)

    time = params.get('time')

    start_time = end_time = None
    if 'today' == time:
        start_time = get_today()
        end_time = get_tomorrow()
    elif 'yestoday' == time:
        start_time = get_yestoday()
        end_time = get_today()
    elif 'toweek' == time:
        start_time = get_first_day_of_week()
        end_time = get_tomorrow()
    elif 'tomonth' == time:
        start_time = get_first_day_of_month()
        end_time = get_tomorrow()
    elif 'other' == time:
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
    if start_time and end_time:
        where.append("a.create_at <= '%s' and a.create_at >= '%s'" % (end_time, start_time))

    if status in ('collection_success', 'partial_success') and start_time and end_time:
        where.append("a.real_repay_time <= '%s' and a.real_repay_time >= '%s'" % (end_time, start_time))

    if params.get("payment_range"):
        pay_min, pay_max = params['payment_range'].split(",")
        try:
            pay_min = int(pay_min)
            pay_max = int(pay_max)
        except:
            pay_min = pay_max = None
        if pay_min:
            where.append('a.rest_repay_money >= %s' % pay_min)
        if pay_max:
            where.append('a.rest_repay_money <= %s' % pay_max)
    if params.get('overdu_days'):
        overdue_days = json.loads(params['overdu_days'])
        start_day = overdue_days.get('start_day')
        end_day = overdue_days.get('end_day')
        if start_day and end_day and end_day >= start_day >= 0:
            where.append("a.overdue_days >= %s and a.overdue_days <= %s" % (start_day, end_day))

    if params.get("search"):
        where.append(
            "b.name like '%{search}%' or b.phone_no='{search}' or b.id_no='{search}' or\
            c.order_number='{search}'".format(search=params['search'].decode("utf-8")))

    return " and ".join(where)


@require_http_methods(['GET'])
def wx_search(request):
    session = request.GET.get('session')
    query = request.GET.get('query')
    employee_id = check_wx_session(session)
    if not employee_id:
        return JsonResponse({
            "code": -1,
            "msg": u"请先登录"
        })
    employee = get_object_or_404(Employee, pk = employee_id)
    apply_list = None
    filter = Q(type__in=['b', 'c', 'd', 'e', 'g', 'h']) & (Q(create_by__name = query) | Q(create_by__phone_no = query) | Q(repayment__order_number = query) | Q(create_by__id_no = query))
    if employee.check_page_permission('/collection/all'):
        apply_list = Apply.objects.filter(filter)
    elif employee.check_page_permission('/collection/mine'):
        apply_list = Apply.objects.filter(Q(employee = employee) & filter)
    else:
        return JsonResponse({
            "code": -1,
            "msg": u"你不是催收员"
        })
    result = []
    for apply in apply_list:
        if not apply.repayment:
            continue
        info = {'apply_id': apply.id,
                'order_number': apply.repayment.order_number,
                'name': apply.create_by.name,
                'channel': apply.create_by.channel}
        result.append(info)
    return JsonResponse({
        'code': 0,
        'total_count': len(result),
        'data': result
    })

@require_http_methods(['GET'])
def wx_apply_info(request):
    def _calc_payment(apply):
        pass
    session = request.GET.get('session')
    employee_id = check_wx_session(session)
    if not employee_id:
        return JsonResponse({
            "code": -1,
            "msg": u"请先登录"
        })
    employee = get_object_or_404(Employee, pk = employee_id)
    apply = Apply.objects.filter(pk = request.GET.get('apply')).first()
    if not apply:
        return JsonResponse({
            'code': -1,
            'msg': u'订单不存在'
        })
    if not employee.check_page_permission('/collection/all') and apply.employee != employee:
        return JsonResponse({
            'code': -1,
            'msg': u'你没有查看该订单的权限'
        })

    status_name = ""
    if apply.status in ['8', '9', 's4']:
        status_name = u'催收完成'
    elif apply.status in ['s1', '0']:
        status_name = u'待分配'
    elif apply.status in ['s3', 'ci']:
        status_name = u'已受理'
    elif apply.status in ['s2', 'i']:
        status_name = u'未受理'
    installment = InstallmentDetailInfo.objects.filter(repayment = apply.repayment, overdue_days__gt = 0).order_by('-installment_number').first()
    repayment_record = DingdangRepaymentRecord.objects.filter(apply = apply, type = 0).order_by('-id').first()
    info = {"apply_id": apply.id,
            "order_number": apply.repayment.order_number,
            "name": apply.create_by.name,
            "phone": apply.create_by.phone_no,
            "id_no": apply.create_by.id_no,
            "status": status_name,
            "installment_number": installment.installment_number if installment else 0,
            "overdue_days": apply.overdue_days,
            "principal": apply.repayment.apply_amount,
            "payment": repayment_record.should_repay_amount if repayment_record else 0,
            "should_pay": apply.rest_repay_money,
            "collection_records": []}
    apply_ids = Apply.objects.filter(repayment = apply.repayment)
    records = CollectionRecord.objects.filter(apply__in = apply_ids).order_by('-create_at')
    for record in records:
        info["collection_records"].append({"type": record.get_record_type_display(),
                                          "promise_time": record.promised_repay_time,
                                          "note": record.collection_note,
                                          "employee": record.create_by.username,
                                          "process_time": record.create_at})
    return JsonResponse({
        "code": 0,
        "data": info
    })

@csrf_exempt
@require_http_methods(['POST'])
def download_bi_table(request):
    """
    前端把bi表格数据传给后台，后台生成excel文件后返回前端
    """
    table_data = request.body
    try:
        table_data = json.loads(table_data)
    except Exception,e:
        return JsonResponse({
            'code': -1,
            'msg': str(e)
        })
    filename = 'bi_%s.xls' % datetime.now().strftime('%y%m%d%H%M%S%f')
    th = table_data['th']
    tbody = table_data['data']
    w = Workbook()
    ws = w.create_sheet('sheet1', 0)
    ws.append(th)
    for data in tbody:
        ws.append(data)
    statistic = table_data['statistic']
    for item in statistic:
        ws.append([item['key'], item['value']])
    w.save(filename)

    try:
        f = open(filename)
        file_data = f.read()
        res = requests.post('%sfile' % settings.IMAGE_SERVER_URL, files = {'file': file_data}, data = {'type': 'xls'})
        result = json.loads(res.content)
        response = JsonResponse({'code': 0, 'url': result['url']})
        os.system('rm %s' % filename)
        return response
    except Exception,e:
        return JsonResponse({'code': -1, 'msg': str(e)})




class CollectionRecordViewSet(CuteViewSet):

    pagination_class = CutePageNumberPagination
    serializer_class = CollectionRecordSerializer
    queryset = CollectionRecord.objects.all()

    def get_queryset(self):
        """queryset 状态 < 0 的不返回"""
        queryset = self.queryset
        if isinstance(queryset, QuerySet):
            queryset = queryset.all()
        if self.platform:
            queryset = queryset.filter(status__gte=0, apply__platform__in = self.platform).order_by('-update_at')
        else:
            queryset = queryset.filter(status__gte=0).order_by('-update_at')
        params = self.request.query_params
        print params
        apply_id = params.get('apply_id')
        # apply = Apply.objects.get(id=apply_id)
        apply = get_object_or_404(Apply, pk=apply_id)

        record_type = params.get('type')
        order_field = params.get('order_way')

        order_fields = ['-create_at']
        if order_field in ['asc']:
            order_fields = ['create_at']

        q = {
            'apply__in': Apply.objects.filter(repayment_id=apply.repayment_id),
        }
        if record_type:
            q.update(dict(record_type=record_type))
        queryset = queryset.filter(**q).order_by(*order_fields)

        return queryset

    def update_apply(self, request, data):
        """更新 apply"""
        pass
        # update apply
        apply_id = request.data['apply_id']
        apply = Apply.objects.get(id=apply_id)

        apply.promised_repay_time = data['promised_repay_time']
        apply.collection_record_time = data['add_time']
        apply_status = apply.status
        if apply_status in ['s2', 'i']:
            apply_status = 's3'
        apply.status = apply_status
        apply.save()

        return apply

    def report_data(self, apply, employee):
        _records = CollectionRecord.objects.filter(apply=apply, create_by=employee, record_type=CollectionRecord.COLLECTION)
        if not _records:
            try:
                report_collection(apply=apply)
            except Exception as e:
                print e

        report_collection(apply=apply, status_name=u'催记')


    def create(self, request, *args, **kwargs):
        # data = get_request_data(request)
        print 'in view_set create'
        data = request.data
        print data
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        apply = self.update_apply(request, serializer.data)
        employee = Employee.objects.filter(user=request.user).first()
        # employee = Employee.objects.filter().first()
        self.report_data(apply, employee)

        data = dict(
            code=self.code,
            msg=self.msg,
            data=serializer.data,
        )

        print 'out view_set create'
        return Response(data, headers=headers)


def format_request_params(check_params, new_args=False):
    """
    parameters: check_params {"arg1": "json", "arg2": int}
    """
    def real_check(func):
        def decorator(*args, **kwargs):
            params = args[0].request.query_params.copy()
            newargs = kwargs
            for param, arg_type in check_params.items():
                if not params.get(param):
                    continue
                if arg_type == "json":
                    try:
                        new_arg = json.loads(params[param])
                    except Exception as e:
                        raise ValueError('json loads error %s' % params[param])
                else:
                    try:
                        new_arg = arg_type(params[param])
                    except Exception as e:
                        raise ValueError("format value " + e)
                newargs[param] = new_arg
            if new_args:
                return func(*args, **newargs)
            else:
                args[0].request.query_params = params.update(newargs)  # TODO
                return func(*args, **kwargs)
        return decorator
    return real_check


class QualityPagination(pagination.PageNumberPagination):

    page_size = 10
    page_size_query_param = 'page_size'

    def get_paginated_response(self, data):
        return JsonResponse({
            'code': data['code'],
            'msg': data['msg'],
            'data': data['data'],
            'count': data['count'],
        })


class CsrfExemptSessionAuthentication(SessionAuthentication):

    def enforce_csrf(self, request):
        return  # To not perform the csrf check previously happening


class QualityControlList(viewsets.ModelViewSet):

    queryset = QualityControlRecord.objects.filter(status=QualityControlRecord.normal)
    serializer_class = QualitySerializer
    pagination_class = QualityPagination

    authentication_classes = (CsrfExemptSessionAuthentication, BasicAuthentication)

    code = 0
    msg = ''

    en_cn_map = {
        u'组名': 'cn_group',
        u'催收员': 'employee',
        u'客户工单号': 'order_number',
        u'客户姓名': 'customer',
        u'客户电话': 'customer_phone',
        u'录音时间': 'recording_time',
        u'质检详情': 'inspection_detail',
        u'质检人': 'quality_people',
        u'处理': 'treatment'
    }

    # @format_request_params({'date': 'json'})
    def get_queryset(self):
        employee = get_object_or_404(Employee, user=self.request.user)
        have_permission = employee.check_page_permission(self.request.path)
        query_set = super(QualityControlList, self).get_queryset()
        params = self.request.query_params
        # print('employee: ', employee, have_permission)
        print params
        # have_permission = True
        if have_permission:   # 管理员页面
            employee_id = params.get('employee_id')
        else:                 # 个人页面
            employee_id = employee.user.username
        query_employee = Q(employee__user__username=employee_id) if employee_id else Q()
        date_ = params.get('date')
        if date_:
            json_date = json.loads(date_)
            start_date = datetime.strptime(json_date.get('start'), "%Y-%m")
            end_date = datetime.strptime(json_date.get('end'), "%Y-%m") + relativedelta(months=1)
            query_date = Q(check_time__gte=start_date)&Q(check_time__lt=end_date)
        else:
            query_date = Q()
        cn_group = params.get('cn_group')
        query_cn_group = Q(cn_group=cn_group) if cn_group else Q()
        query_issue = Q()
        if params.get('issues'):
            query_issue = Q() if int(params['issues']) == 0 else ~Q(inspection_detail__warn_level=0)
        ret_queryset = query_set.filter(query_employee & query_date
                                        & query_cn_group & query_issue)
        return ret_queryset

    def list(self, request, *args, **kwargs):
        """
        HTTP METHOD: GET
        """
        queryset = self.filter_queryset(self.get_queryset())
        count = self.get_queryset().count()
        ret = {
            "code": self.code,
            "msg": self.msg,
            "data": [],
            "count": count
        }
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            ret['data'] = serializer.data
            return self.get_paginated_response(ret)

        serializer = self.get_serializer(queryset, many=True)
        ret['data'] = serializer.data
        return JsonResponse(ret)

    def create(self, request):
        """
        HTTP METHOD: POST
        """
        upload_file = request.FILES
        params = request.POST.dict()
        if not upload_file.get('file'):
            return JsonResponse({
                "code": -1,
                "msg": "params error"
            })
        if params.get('date'):
            date_ = datetime.strptime(params['date'], "%Y-%m")
        else:
            date_ = datetime.now()
        import_file = ImportFile(upload_file['file'])
        header = import_file.header
        header = header[:9]
        content = import_file.sheet_content
        data = []
        for item in content:
            item = item[:9]
            if not any(item):
                continue
            data.append(item)
        en_data = self.convert_data(header, data[1:])
        ret = QualityControlRecord.import_record(date_, en_data)
        ret = {
            "code": self.code,
            "msg": self.msg,
            "data": ret,
        }
        return JsonResponse(ret)

    def destroy(self, request, *args, **kwargs):
        """
        HTTP METHOD: DELETE
        """
        record = self.get_object()
        if not record:
            return JsonResponse({
                "code": -1,
                "msg": "record_id is null"
            })
        record.delete()
        return JsonResponse({
            "code": 0,
            "msg": "",
            "data": ""
        })

    @detail_route(methods=['post'])
    def fake_delete(self, request, pk=None):
        """
        url: quality_control/{id}/fake_delete/   method: post
        """
        instance = self.get_object()
        if instance.status == QualityControlRecord.deleted:
            return JsonResponse({
                "code": -1,
                "msg": "不存在",
                "data": ""
            })
        method = request.POST.get('method')
        if method != 'delete':
            return JsonResponse({
                "code": -1,
                "msg": "参数错误",
                "data": ""
            })
        instance.status = QualityControlRecord.deleted
        instance.save()
        return JsonResponse({
            "code": 0,
            "msg": "",
            "data": ""
        })

    def convert_data(self, header, content):
        # need to convert Chinese to English
        # return dict
        ret = []
        en_index = map(lambda item: self.en_cn_map.get(item), header)
        for line in content:
            ret.append({en_index[i]: j for i, j in enumerate(line)})
        return ret


class CnGroupView(generics.GenericAPIView):

    code = 0
    msg = ''

    def get(self, request, *args, **kwargs):
        cn_groups = QualityControlRecord.objects.filter(status=QualityControlRecord.normal).values_list('cn_group').distinct()
        return JsonResponse({
            "code": self.code,
            "msg": self.msg,
            "data": [item[0] for item in cn_groups]
        })


class InspectionResult(generics.GenericAPIView):

    queryset = QualityControlRecord.objects.all()
    code = 0
    msg = ''

    def get(self, request, *args, **kwargs):
        """
        django  1.7 不支持case when, 所以用sql
        """
        where = self.get_sql_filter(request) + "and status = %s " % QualityControlRecord.normal
        sql = "SELECT id, cn_group, check_time, employee_id, \
        sum(case when treatment='%s' then 1 else 0 end) as warn, \
        sum(case when treatment='%s' then 1 else 0 end) as work_lost, \
        sum(case when treatment='%s' then 1 else 0 end) as offline\
        from quality_control_record %s\
        GROUP BY employee_id" % (InspectionDetails.warn, InspectionDetails.lost_work, InspectionDetails.offline, where)
        print sql
        ret = self.queryset.raw(sql)
        data = []
        for item in ret:
            data.append({
                "employee": item.employee.username,
                "cn_group": item.cn_group,
                "month": item.check_time.strftime("%m"),
                "warn": int(item.warn),
                "lost": int(item.work_lost),
                "offline": int(item.offline)
            })
        return JsonResponse({
            "code": self.code,
            "msg": self.msg,
            "data": data
        })

    def get_sql_filter(self, request):
        sql_where = []
        params = request.GET.dict()
        employee = get_object_or_404(Employee, user=self.request.user)
        have_permission = employee.check_page_permission(self.request.path)
        # have_permission = True
        if have_permission:   # 管理员页面
            if params.get('employee_id'):
                staff = Employee.objects.filter(user__username=params['employee_id']).first()
                if staff:
                    sql_where.append("employee_id = %s" % staff.id)
        else:
            sql_where.append("employee_id = %s" % employee.id)
        date_ = params.get('date')
        cn_group = params.get('cn_group')
        if date_:
            date_dict = json.loads(date_)
            start_date = datetime.strptime(date_dict.get('start'), "%Y-%m")
            end_date = datetime.strptime(date_dict.get('end'), "%Y-%m") + relativedelta(months=1)
            sql_where.append("check_time >= '%s' and check_time < '%s'" % (start_date, end_date))
        else:
            date_cond = datetime.now()
            sql_where.append("date_format(check_time, '%%Y%%m') = '{}'".format(date_cond))
        if cn_group:
            sql_where.append("cn_group = '%s'" % cn_group.strip().decode("utf-8"))
        if not sql_where:
            return ""
        elif len(sql_where) == 1:
            return " where " + sql_where[0]
        else:
            return " where " + " and ".join(sql_where)


class InspectionAnalysis(generics.GenericAPIView):

    queryset = QualityControlRecord.objects.filter(status=QualityControlRecord.normal)
    code = 0
    msg = ''

    # @format_request_params({'date': 'json'})
    def get_queryset(self):
        params = self.request.query_params
        date_ = params.get('date')
        cn_group = params.get('cn_group')
        employee = get_object_or_404(Employee, user=self.request.user)
        have_permission = employee.check_page_permission(self.request.path)
        # have_permission = True
        if have_permission:   # 管理员页面
            employee_id = params.get('employee_id')
        else:                 # 个人页面
            employee_id = employee.id
        query_employee = Q(employee_id=employee_id) if employee_id else Q()
        if date_:
            date_dict = json.loads(date_)
            start_date = datetime.strptime(date_dict.get('start'), "%Y-%m")
            end_date = datetime.strptime(date_dict.get('end'), "%Y-%m") + relativedelta(months=1)
            query_date = Q(check_time__gte=start_date)&Q(check_time__lt=end_date)                # 这里的range是sql里的between
        else:
            date_cond = datetime.now()
            query_date = Q(check_time__year=date_cond.year, check_time__month=date_cond.month)
        queryset = super(InspectionAnalysis, self).get_queryset()
        query_group = Q(cn_group=cn_group.strip()) if cn_group else Q()
        return queryset.filter(query_group & query_date & query_employee)

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        rets = queryset.values('inspection_detail__cn_name').annotate(icount=Count('inspection_detail'))
        data = [
            {
                'value': item['icount'],
                'name': item['inspection_detail__cn_name']
            }
            for item in rets
        ]
        return JsonResponse({
            'code': self.code,
            'msg': self.msg,
            'data': data
        })


@require_http_methods(["POST"])
@csrf_exempt
def lost_contact(request):
    """
    催收失联
    """
    params = json.loads(request.body)
    order_id = params.get('order_id')
    # check status
    employee = get_object_or_404(Employee, user_id=request.user.id)
    app = get_object_or_404(Apply, pk=order_id)
    # if app.status not in (Apply.PROCESSING, Apply.APPLY_PROCESSING):
    #     # not except apply status
    #     return JsonResponse({
    #         "code": -1,
    #         "msg": u"状态不对",
    #         "data": ""
    #     })
    # change apply status
    note = u'{} 标记失联'.format(employee.username)
    app.status = Apply.LOST_CONTACT
    app.save()
    # create collection record
    collection_record = dict(
        record_type=CollectionRecord.LOST_CONTACT,
        collection_note=note,
        create_by=employee,
        apply=app
    )
    CollectionRecord.objects.create(**collection_record)
    report_collection(app, app.get_status_display())
    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": ""
    })


@require_http_methods(["POST"])
@csrf_exempt
def verify_lost_order(request):
    """
    信修核实
    """
    msg_temp = u"核实操作: {}"
    params = json.loads(request.body)
    order_id = params.get('order_id')
    verify_ret = params.get('verify_ret')
    employee = get_object_or_404(Employee, user_id=request.user.id)
    app = get_object_or_404(Apply, pk=order_id)
    records = InfoField.objects.filter(status=InfoField.VALID, is_must=True, user=app.create_by)
    change_records = InfoField.objects.filter(                     # 有效和未验证
        user=app.create_by, status__in=(InfoField.VALID, InfoField.NOT_VERIFY), is_must=True
    )
    success_records = InfoField.objects.filter(
        user=app.create_by, status=InfoField.VALID, is_must=True
    )
    not_verify_records = InfoField.objects.filter(
        user=app.create_by, status=InfoField.NOT_VERIFY
    )
    if int(verify_ret) and records:
        code = 0
        app.status = Apply.WAIT_DISTRIBUTION
        msg = u"信修成功"
        note = msg_temp.format(msg)
        app.extra_info = Apply.INFO_REPAIR
    elif not success_records and not_verify_records:
        code = -1
        msg = u"订单不可进行状态扭转"
        note = msg_temp.format(msg)
    elif not change_records:           # 全为无效,没有有效和未验证
        app.extra_info = Apply.INFO_REPAIR
        code = -1
        app.status = Apply.REPAIR_FAIL
        msg = u"信修失败"
        note = msg_temp.format(msg)
    app.save()
    collection_record = dict(
        record_type=CollectionRecord.LOST_CONTACT,
        collection_note=note,
        create_by=employee,
        apply=app
    )
    CollectionRecord.objects.create(**collection_record)
    report_collection(app, app.get_status_display())
    return JsonResponse({
        "code": code,
        "msg": msg,
        "data": ""
    })


@require_http_methods(["POST"])
@csrf_exempt
def verify_info_field(request):
    """
    核实信修字段
    """
    params = json.loads(request.body)
    field_id = params.get('id')
    verify_ret = params.get('verify_ret')
    employee = get_object_or_404(Employee, user_id=request.user.id)
    field = get_object_or_404(InfoField, pk=field_id)
    if int(verify_ret) == InfoField.INVALID:
        field.status = InfoField.INVALID
    elif int(verify_ret) == InfoField.VALID:
        field.status = InfoField.VALID
    elif int(verify_ret) == InfoField.NOT_VERIFY:
        field.status = InfoField.NOT_VERIFY
    else:
        return JsonResponse({
            "code": -1,
            "msg": u"错误状态"
        })
    field.save()
    RepairHistory.objects.create(
        info_field=field, employee=employee, operation_status=field.status
    )
    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": ""
    })


@require_http_methods(["GET"])
def info_repair_history(request):
    """
    历史
    """
    field_id = request.GET.get("id")
    histories = RepairHistory.objects.filter(info_field_id=field_id).order_by("-operation_time")
    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": [
            {
                "date": item.operation_time.strftime("%Y-%m-%d %H:%M:%S"),
                "employee": item.employee.username,
                "status": item.operation_status
            } for item in histories
        ]
    })


@require_http_methods(['GET'])
def get_info_repair_module(request):
    """
    信修模块
    """
    order_id = request.GET.get('order_id')
    app = get_object_or_404(Apply, pk=order_id)
    info_modules = InfoField.objects.filter(user=app.create_by).values_list('info_module').distinct()
    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": [{'cn_name': item[0]} for item in info_modules]
    })


@require_http_methods(['GET'])
def info_repair_detail(request):
    """
    信修详情
    """
    params = request.GET.dict()
    order_id = params.get('order_id')
    page_size = params['page_size'] if params.get('page_size') else 10
    app = get_object_or_404(Apply, pk=order_id)
    info_fields = InfoField.objects.filter(
        user=app.create_by, info_module=params['info_module']
    ).order_by("cn_name")
    paginator = Paginator(info_fields, page_size)
    data = paginator.page(params['page'])
    return JsonResponse({
        "code": 0,
        "msg": "",
        "is_repaired": 1 if app.extra_info == 'a' else 0,
        "count": info_fields.count(),
        "data": [
            {
                "id": item.id,
                "info_category": item.cn_name,
                "info": item.content,
                "status": item.status
            } for item in data.object_list
        ]
    })


@require_http_methods(['GET'])
def apply_status(request):
    data = [
        ['wait_distribution', u'待分配'],
        ['not_accept', u'未受理'],
        ['processing', u'已受理'],
        ['wait_check', u'待复核'],
        ['check_failed', u'复核失败'],
        ['collection_success', u'催收完成'],
        ['partial_success', u'部分成功'],
        ['repay_failed', u'扣款失败'],
        ['repay_error', u'扣款异常'],
        ['lost_contact', u'已失联'],
        ['recall_fail', u'召回失败'],
        ['recall_success', u'召回成功'],
        ['repair_fail', u'信修失败'],
        ['renew', u'续期'],
    ]
    data = OrderedDict(data)
    print data

    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": data
    })


@require_http_methods(['GET'])
def my_apply_status(request):
    data = [
        ['wait_distribution', u'待分配'],
        ['not_accept', u'未受理'],
        ['processing', u'已受理'],
        ['wait_check', u'待复核'],
        ['check_failed', u'复核失败'],
        ['collection_success', u'催收完成'],
        ['partial_success', u'部分成功'],
        ['repay_failed', u'扣款失败'],
        ['repay_error', u'扣款异常'],
        ['renew', u'续期'],
    ]
    data = OrderedDict(data)
    print data

    return JsonResponse({
        "code": 0,
        "msg": "",
        "data": data
    })



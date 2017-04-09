# -*- coding:utf-8 -*-
import csv
import os
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from xlrd import open_workbook, xldate_as_tuple
from django.core.files import File

from business_manager.import_data.models import ImportField, ImportModule
from business_manager.import_data.models import ParseFileRecord


class ParseCsvFile(object):

    def __init__(self, file_):
        self.file_ = file_
        self.gbk_count = 0
        self.utf_count = 0
        self.spamreader = [line for line in self.read_csv()]

    def read_csv(self):
        try:
            reader = csv.reader(BytesIO(self.file_.read()))
            return reader
        except Exception as e:
            print("read_csv error: %s" % e)
            raise ImportFileException(msg="read csv error %s" % e)

    def get_header_name(self):
        """
        get first line content
        :return: list
        """
        for row in self.spamreader:
            ret = []
            for cell in row:
                code = self.get_code(cell)
                if code:
                    ret.append(cell.decode(code))
                else:
                    raise ImportFileException("not support file encoding")
            return ret

    def get_sheet_content(self):
        """
        get sheet content
        :return: list or dict ?
        """
        content = []
        print "in get_sheet_content"
        for line in self.spamreader:
            ret = []
            for cell in line:
                if self.gbk_count > 100:
                    code = 'gbk'
                elif self.utf_count > 100:
                    code = 'utf-8'
                else:
                    code = self.get_code(line[0])
                if code:
                    print cell, code, type(cell)
                    ret.append(cell.decode(code))
                else:
                    raise ImportFileException("not support file encoding")
            content.append(ret)
            # 当有一行为空时, 停止读取
            if not any(ret):
                break;

        print len(content)
        print "out get_sheet_content"
        return content

    def get_code(self, content):
        try:
            content.decode('utf-8')
            print content.decode('utf-8')
            print("content %s" % content)
        except UnicodeDecodeError:
            try:
                content.decode('gbk')
            except UnicodeDecodeError:
                return None
            else:
                self.gbk_count += 1
                return 'gbk'
        else:
            self.utf_count += 1
            return 'utf-8'


class ParseXlsxFile(object):

    def __init__(self, file_):
        self.file_ = file_
        self.workbook = self.read_xlsx()

    def read_xlsx(self):
        print 'ParseXlsxFile: read_xlsx start'
        try:
            return load_workbook(filename=BytesIO(self.file_.read()), data_only=True)
        except Exception as e:
            print("read xlsx error: %s" % e)
            raise ImportFileException(msg='read xlsx error: %s' % e)
        print 'ParseXlsxFile: read_xlsx end'

    def get_header_name(self):
        """
        get first line content
        :return: list
        """
        for wk in self.workbook:
            for line in wk:
                if line[0].value:   # will return [None]
                    return [l.value for l in line]
                else:
                    return []

    def get_sheet_content(self):
        """
        get sheet content
        :return: list or dict ?
        """
        content = []
        for wk in self.workbook:
            print wk
            for line in wk:
                _d = [l.value for l in line]
                content.append(_d)

                # 当有一行为空时, 停止读取
                if not any(_d):
                    break;

            # print "in xlsx", content
            print len(content)

            return content

    @staticmethod
    def virtual_save(content, header_mapping):
        print("header_mapping: ",  header_mapping)
        ret_wb = Workbook()
        ret_sheet = ret_wb.create_sheet("fail_data", 0)
        error = header_mapping.pop('error')
        mapping = header_mapping.items()
        mapping.append((error, error))    # 为了把error写在最后一行
        first_line = [item[1] for item in mapping]
        line_index = [item[0] for item in mapping]
        for col, data in enumerate(first_line):
            ret_sheet.cell(column=col+1, row=1, value=data)
        for row, item in enumerate(content):
            for data in item:
                col = line_index.index(data)
                ret_sheet.cell(column=col+1, row=row+2, value=item[data])
        return save_virtual_workbook(ret_wb)


class ParseXlsFile(object):

    def __init__(self, file_):
        self.file_ = file_
        self.workbook = self.read_xls()
        self.header_len = len(self.get_header_name())

    def read_xls(self):
        try:
            return open_workbook(file_contents=self.file_.read())
        except ImportFileException as e:
            print("read xls error: %s" % e)
            raise ImportFileException(msg='read xlsx error: %s' % e)

    def get_header_name(self):
        """
        get first line content
        :return: list
        """
        first_line = []
        first_sheet = self.workbook.sheets()[0]
        header = first_sheet.row_values(0)
        if header:
            first_line.extend(header)
        return first_line

    def get_sheet_content(self):
        """
        get sheet content
        :return: list or dict ?
        """
        content = []  # TODO parse xls content
        first_sheet = self.workbook.sheets()[0]
        nrows = first_sheet.nrows
        for col in range(nrows):
            line = first_sheet.row_values(col)
            if len(line) < self.header_len:
                line.extent([""] * (self.header_len - len(line)))
            content.append(line)

            # 当有一行为空时, 停止读取
            if not any(line):
                break;
        # print ("in xls: ", content)
        print len(content)
        return content


class ImportFileException(BaseException):

    def __init__(self, msg):
        self.msg = msg


class ImportFile(object):

    FILE_TYPE_MAP = {'csv': ParseCsvFile,
                     'xls': ParseXlsFile,
                     'xlsx': ParseXlsxFile}

    def __init__(self, file_):
        self.file_ = file_

        self.file_type = self.get_file_type()
        self.parse_class = self.get_parse_class()
        self.header_mapping = {}

    @property
    def header(self):
        try:
            header = self.parse_class.get_header_name()
        except Exception as e:
            print("get header error: %s" % e)
            raise ImportFileException(msg="get header error: %s" % e)
        return header

    @property
    def sheet_content(self):
        try:
            sheet_content = self.parse_class.get_sheet_content()
        except Exception as e:
            print("get sheet content error: %s" % e)
            raise ImportFileException(msg="get sheet content error :%s" % e)
        return sheet_content

    def get_parse_class(self):
        """
        get parse class
        :return:
        """
        parse_cls = self.FILE_TYPE_MAP.get(self.file_type)
        if not parse_cls:
            # not support file type
            raise ImportFileException(msg="not support file type")
        return parse_cls(self.file_)

    def get_file_type(self):
        """
        get file type
        :return:
        """
        file_type = os.path.splitext(self.file_.name)[-1]
        file_type = file_type[1:] if len(file_type) > 1 else ''
        return file_type

    def process_content(self, record_id):
        # 1. check header
        # 2. process context
        print 'out parse_content'
        ret = []
        en_name_header = []
        record = ParseFileRecord.objects.filter(id=record_id).first()
        module = record.module
        header = self.header
        field_mappings = ImportField.objects.filter(module=module, status__gte=0)
        print("process_content: ", header, field_mappings)
        header_set = set(item.strip() if item else '' for item in header)
        print '-----------'
        sys_set = set(item.user_field_name.strip() for item in field_mappings)
        if sys_set - header_set:
            print sys_set - header_set
            record.status = ParseFileRecord.HEADER_MATCH_ERROR
            record.save()
            raise ImportFileException(msg="header match error")
        for cn_name in header:
            if not cn_name:
                continue
            for field in field_mappings:
                if field.user_field_name.strip() == cn_name.strip():
                    self.header_mapping[field.sys_field_id.name] = cn_name
                    en_name_header.append((field.sys_field_id.name, header.index(cn_name)))
        # for cn_name, field_mapping in zip(header, field_mappings):
        #     print("header match %s, %s" % (field_mapping.user_field_name, cn_name))
        #     if field_mapping.user_field_name.strip() == cn_name.strip():
        #         self.header_mapping[field_mapping.sys_field_id.name] = cn_name
        #         en_name_header.append(field_mapping.sys_field_id.name)
        #     else:
        #         # header match failed
        #         record.status = ParseFileRecord.HEADER_MATCH_ERROR
        #         record.save()
        #         raise ImportFileException(msg="header match error")
        print 'out parse_content'
        self.header_mapping['error'] = 'error'
        content = self.sheet_content[1:]
        for line in content:
            dict_line = {}
            if not any(line):
                continue
            for item in en_name_header:
                value = line[item[1]]
                # if "time" in item[0] and isinstance(line[item[1]], float):
                    # print("convert xls time ", item, value)
                    # try:
                        # value = datetime(*xldate_as_tuple(value, 0)).strftime("%Y-%m-%d")
                    # except Exception as e:
                        # print("convert xls time error", e)
                        # value = ""
                dict_line[item[0]] = value
            # ret.append(dict(zip(en_name_header, line)))
            ret.append(dict_line)
        return ret

    def virtual_save(self, content, record_id):
        """
        default save xlsx
        :param record_id:
        :param content: data list, first is the header
        :return: str
        """
        ret = ParseXlsxFile.virtual_save(content, self.header_mapping)
        return FailFile(record_id, ret)

    @classmethod
    def save_fail_file(cls, content, record_id, en_cn_map):
        ret = ParseXlsxFile.virtual_save(content, en_cn_map)
        file_name = "fail_data_{}.xlsx".format(record_id)
        fail_file = File(BytesIO(ret), file_name)
        return file_name, fail_file


class FailFile:

    def __init__(self, record_id, content):
        self.file_name = "fail_data_{}.xlsx".format(record_id)
        self.file_content = File(BytesIO(content), self.file_name)
